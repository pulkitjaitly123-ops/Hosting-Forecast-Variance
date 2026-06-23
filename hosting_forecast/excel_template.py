"""
Build a sample Excel upload template for the dashboard.

The dashboard's loader (`data.load_actuals`) reads the FIRST worksheet, so the
data lives on sheet 1 ("Hosting Actuals") and the guidance lives on sheet 2
("Instructions"). The file doubles as a worked example and a blank-able template:
replace the sample rows with your own monthly actuals and upload it as-is.

Run:
    python -m hosting_forecast.excel_template
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import config
from .data import load_actuals

TEMPLATE_PATH = os.path.join(config.SAMPLE_DIR, "hosting_actuals_template.xlsx")

# Palette (hex without '#', matching build_model.py conventions)
NAVY = "1B3A6B"
TEAL = "00A4A6"
LGREY = "F3F4F6"
BLUE_FC = "0000FF"   # input cells

# column -> (header, number format, required?, description)
COLUMNS = [
    ("month",             "yyyy-mm-dd", True,
     "First day of the month, e.g. 2025-01-01"),
    ("opening_customers", "#,##0",      True,
     "Hosting customers at the start of the month"),
    ("new_customers",     "#,##0",      True,
     "Gross new customers added during the month"),
    ("churned_customers", "#,##0",      True,
     "Customers lost during the month"),
    ("closing_customers", "#,##0",      True,
     "Customers at month end (opening + new - churned)"),
    ("arpc",              "$#,##0.00",  True,
     "Average revenue per customer for the month ($/customer)"),
    ("revenue",           "$#,##0",     True,
     "Monthly revenue (average customers x ARPC)"),
    ("budget_revenue",    "$#,##0",     False,
     "Optional: planned/budget revenue for the month"),
    ("budget_customers",  "#,##0",      False,
     "Optional: planned closing customers"),
    ("budget_arpc",       "$#,##0.00",  False,
     "Optional: planned ARPC"),
]

_thin = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_cell(c, text):
    c.value = text
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border


def build_template(path: str = TEMPLATE_PATH) -> str:
    """Write the sample/template workbook and return its path."""
    config.ensure_dirs()
    df = load_actuals()  # bundled sample (generates it if missing)
    headers = [c[0] for c in COLUMNS]

    wb = Workbook()

    # ── Sheet 1: data (what the loader reads) ──────────────────────
    ws = wb.active
    ws.title = "Hosting Actuals"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for j, (name, fmt, required, _desc) in enumerate(COLUMNS, start=1):
        _header_cell(ws.cell(row=1, column=j), name)

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, (name, fmt, required, _desc) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=i, column=j)
            val = row.get(name)
            if name == "month":
                cell.value = row["month"].to_pydatetime()
            else:
                cell.value = None if val is None else float(val)
            cell.number_format = fmt
            cell.font = Font(name="Arial", color=BLUE_FC, size=10)
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="left" if name == "month" else "right")

    for j, (name, *_x) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(name) + 3)
    ws.row_dimensions[1].height = 20

    # ── Sheet 2: instructions ──────────────────────────────────────
    info = wb.create_sheet("Instructions")
    info.sheet_view.showGridLines = False
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 12
    info.column_dimensions["C"].width = 14
    info.column_dimensions["D"].width = 60

    title = info.cell(row=1, column=1, value="How to use this template")
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = PatternFill("solid", fgColor=NAVY)
    info.merge_cells("A1:D1")
    info.row_dimensions[1].height = 24

    notes = [
        "Put one row per month on the 'Hosting Actuals' sheet (sheet 1).",
        "Keep the header names exactly as shown; the loader reads sheet 1.",
        "Required columns must be filled. Budget columns are optional: leave",
        "   them out (or blank for all rows) to get a forecast with no variance.",
        "Upload the saved .xlsx (or a CSV with the same columns) in the app sidebar.",
    ]
    r = 3
    for n in notes:
        info.cell(row=r, column=1, value=("- " + n)).font = Font(name="Arial", size=10)
        info.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    r += 1
    for j, h in enumerate(["Column", "Type", "Required", "Description"], start=1):
        _header_cell(info.cell(row=r, column=j), h)
    r += 1
    for name, fmt, required, desc in COLUMNS:
        info.cell(row=r, column=1, value=name).font = Font(name="Arial", size=10)
        info.cell(row=r, column=2,
                  value="date" if name == "month" else "number").font = Font(name="Arial", size=10)
        rc = info.cell(row=r, column=3, value="Yes" if required else "Optional")
        rc.font = Font(name="Arial", size=10,
                       color="1F6B2E" if required else "888888")
        info.cell(row=r, column=4, value=desc).font = Font(name="Arial", size=10)
        r += 1

    wb.save(path)
    return path


if __name__ == "__main__":
    p = build_template()
    print(f"Wrote template -> {p} ({os.path.getsize(p):,} bytes)")
    # round-trip check: the loader must accept it
    df = load_actuals(p)
    print(f"Loader read it back: {len(df)} rows, columns OK, "
          f"latest revenue ${df['revenue'].iloc[-1]:,.0f}")
